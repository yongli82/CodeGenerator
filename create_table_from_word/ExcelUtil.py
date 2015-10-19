import os
import sys
reload(sys)
sys.path.append("..")
sys.setdefaultencoding('utf-8')

from openpyxl import load_workbook

def cell_value(cell, default=""):
    value = cell.value
    if value:
        if isinstance(value, str):
            return value.strip()
        else:
            return value
    else:
        return default

def read_cell(ws, row, column):
    cell = ws['%s%s' % (column, row)]
    return cell_value(cell)

def read_cells(ws, row, columns):
    values = []
    for column in columns:
        values.append(read_cell(ws, row, column))

    return values

def read_excel(file_path, sheet_name, columns):
    wb = load_workbook(file_path)
    ws = wb.get_sheet_by_name(sheet_name)
    BEGIN_ROW = 1
    END_ROW = len(ws.rows)
    data_list = []
    for row in range(BEGIN_ROW, END_ROW + 1):
        values = read_cells(ws, row, columns)
        isEmpty = True
        for value in values:
            if not isinstance(value, str) or len(value) > 0:
                isEmpty = False

        if not isEmpty:
            data_list.append(values)

    return data_list


def read_excel_with_head(file_path, sheet_name, columns):
    data_list = read_excel(file_path, sheet_name, columns)
    head = data_list[0]
    data_grid = []
    for row in data_list[1:]:
        record = {}
        for i in range(0, len(head)):
            title = head[i]
            value = row[i]
            record[title] = value
        data_grid.append(record)

    return data_grid

#TODO
def calc_column_num(last):
    letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    if last is None or len(last) == 0:
        return 'A'
    num = 0
    digital = 1
    last = last.upper()
    lastlist = list(last)
    lastlist.reverse()
    for letter in lastlist:
        num += (letters.index(letter) + 1) * digital
        digital *= 26
    return num

def calc_column_name(num):
    letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    nlist = []
    while num > 0:
        mod = num % 26
        num = num / 26
        if mod == 0:
            nlist.append(letters[26 - 1])
            num = num - 1
        else:
            nlist.append(letters[mod - 1])
    nlist.reverse()
    return "".join(nlist)


def generate_columns(start, end):
    start_num = calc_column_num(start)
    end_num = calc_column_num(end)
    columns = []
    for i in range(start_num, end_num):
        columns.append(calc_column_name(i))
    return columns


if __name__ == "__main__":
    columns = generate_columns("A", "AAA")
    print columns
